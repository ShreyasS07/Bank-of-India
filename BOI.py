import os
import time
import psutil
import openpyxl
import threading
import pandas as pd
import tkinter as tk
from bs4 import BeautifulSoup
from selenium import webdriver
from tkinter.messagebox import askyesno
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException

loopCount = 0
runTimer = 3600  # in sec
treadLoop = None

output_folder = "Output Files"
if not os.path.exists(output_folder):
    os.mkdir(output_folder)

master_file = os.path.join(output_folder, "Master.xlsx")
slave_file = os.path.join(output_folder, "Slave.xlsx")

def forever():
    global loopCount, treadLoop, text2show, output_folder, master_file, slave_file
    treadLoop = threading.Timer(runTimer, forever)
    treadLoop.start()
    loopCount += 1
    print("\nProcess-",loopCount)
    # service = Service("C:\\Users\\ASUS\\Downloads\\SBI\\chromedriver_win32\\chromedriver.exe")
    service = Service("C:\\Users\\ASUS\\PycharmProjects\\Bank of India\\chromedriver_win32\\chromedriver.exe")
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    driver = webdriver.Chrome(service=service, options=options)

    # Webpage Link
    url = 'https://www.ccilindia.com/IRSS_HOME.aspx'
    try:
        driver.get(url)
        time.sleep(2)
    except TimeoutException:
        print("TimeoutException: Failed to load the Webpage")
        # driver.quit()

    # Extracting the HTML content
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')

    # Extracting the 1st table
    # table = soup.select_one('table')
    table = soup.select('table')[7]
    df = pd.read_html(str(table))[0]
    # driver.quit()

    new_columns = ['Tenor', 'Last Reported Rate of Prev day', 'Open', 'High', 'Low', 'Wt. Avg. Rate',
                   'Last Reported Rate of Current day', 'Last Reported Trade Time Stamp for the Current day', 'Nan',
                   'Volume (Crs.)', 'No. of Trades']
    df.columns = new_columns
    df = df.drop(0)
    df = df.drop(df.index[-1])

    # To add Empty Row in between Tables
    # empty_df = pd.DataFrame(columns=df.columns)

    try:
        # master_df = pd.read_excel('Master.xlsx')
        master_df = pd.read_excel("C:\\Users\\ASUS\\PycharmProjects\\Bank of India\\Output Files\\Master.xlsx")
    except FileNotFoundError:
        print("Old Master file not found creating the New file")
        master_df = pd.DataFrame()
    master_df = pd.concat([master_df, df], axis=0)
    # master_df.to_excel('Master.xlsx', index=False)
    master_df.to_excel(master_file, index=False)
    print("Master excel file Saved.")

    # Slave Dtaframe
    data = [['INR', 37, 1, 4], ['INR', 37, 2, 4], ['INR', 37, 3, 4], ['INR', 37, 6, 4], ['INR', 37, 9, 4],
            ['INR', 37, 1, 8], ['INR', 37, 2, 8], ['INR', 37, 3, 8], ['INR', 37, 4, 8], ['INR', 37, 5, 8],
            ['INR', 37, 7, 8], ['INR', 37, 10, 8]]

    slave_columns = ['Col1', 'Col2', 'Col3', 'Col4']
    slave = pd.DataFrame(data, columns=slave_columns)
    df = df.reset_index(drop=True)
    slave['DF_Data'] = df['Last Reported Rate of Current day']

    try:
        # slave_wb = openpyxl.load_workbook('Slave.xlsx')
        slave_wb = openpyxl.load_workbook("C:\\Users\\ASUS\\PycharmProjects\\Bank of India\\Output Files\\Slave.xlsx")
    except FileNotFoundError:
        print("Old Slave file not found creating the New file")
        slave_wb = openpyxl.Workbook()
    slave_ws = slave_wb.active
    for row in slave.itertuples(index=False):
        slave_ws.append(row)
    # slave_wb.save('Slave.xlsx')
    slave_wb.save(slave_file)
    print('Slave Excel file Saved.')
    print("Continuing the Process after 1 hour. ")

def destroy_me():
    global window, treadLoop
    answer = askyesno(title='Mindful Automation Pvt Ltd', message='Are you sure you want to Quit ?')
    if (answer):
        try:
           treadLoop.cancel()
        except:
            pp = 0
        current_system_pid = os.getpid()
        ThisSystem = psutil.Process(current_system_pid)
        ThisSystem.terminate()
        window.destroy()

root = tk.Tk()
root.title("Mindful Automation Pvt Ltd")
root.geometry("350x150")
root['bg'] = 'white'
current_path = os.path.dirname(os.path.realpath(__file__))
root.wm_iconbitmap(f"{current_path}/icons/mindful_logo.ico")
label_1 = tk.Label(root, text="Bank Of India Process-1", width=50, height=3, fg="#03001C")
label_1.grid(column=1, row=1)
start = tk.Button(root, text="Start Process", command=forever, height=1, width=32)
start.grid(column=1, row=2, pady=10)
end = tk.Button(root, text="End Process", command=destroy_me, height=1, width=32)
end.grid(column=1, row=3, pady=10)
root.mainloop()